"""
Convierte el reporte de facturación del sistema al formato Excel de ARCA.
Uso: py scripts/convertir_reporte.py reporte.xlsx [mes] [anio] [salida.xlsx]
"""
import sys
import re
from collections import defaultdict
import openpyxl
from openpyxl import Workbook

MESES = ["", "ENERO", "FEBRERO", "MARZO", "ABRIL", "MAYO", "JUNIO",
         "JULIO", "AGOSTO", "SEPTIEMBRE", "OCTUBRE", "NOVIEMBRE", "DICIEMBRE"]

# Columnas de destino en el Excel ARCA (1-indexed)
# Bloque fijo
COL_NOMBRE   = 1   # A
COL_TIPO     = 2   # B
COL_DIR      = 4   # D
COL_TOTAL    = 5   # E  (fórmula)
# Item 1: col F(6), G(7), H(8), I(9)
# Item 2: col J(10), K(11), L(12), M(13)
# Items 3-9: base cols N(14), R(18), V(22), Z(26), AD(30), AH(34), AL(38)
# Items 3+: base col = 14 + i*4 (4 columnas por ítem, sin límite)

def nombre_display(cliente_raw: str) -> str:
    """'APELLIDO, NOMBRE' → 'NOMBRE APELLIDO'. Sin coma: duplica."""
    if "," in cliente_raw:
        partes = [p.strip() for p in cliente_raw.split(",", 1)]
        return f"{partes[1]} {partes[0]}"
    return f"{cliente_raw} {cliente_raw}"

def aplicar_iva(importe: float, desc: str) -> float:
    if re.search(r"canon|cobrador", desc, re.IGNORECASE):
        return importe
    return round(importe * 1.21, 2)

def extraer_sv(codigo_alfa: str) -> str:
    """'SV-3854/0' → '3854'"""
    m = re.search(r"SV-([^/]+)", str(codigo_alfa), re.IGNORECASE)
    return m.group(1) if m else codigo_alfa

def leer_reporte(path: str) -> dict:
    """Lee el reporte fuente y agrupa las filas por ORDER_ID."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    col = {h: i for i, h in enumerate(headers)}

    grupos = defaultdict(list)
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[col["ORDER_ID"]]:
            continue
        oid = row[col["ORDER_ID"]]
        grupos[oid].append(row)
    return grupos, col

def procesar_cliente(filas: list, col: dict, mes: int, anio: int, direccion: str = "") -> dict:
    """Transforma las filas de un cliente al dict que escribe en ARCA."""
    primera = filas[0]
    cliente_raw = str(primera[col["CLIENTE"]])
    codigo_alfa  = str(primera[col["CODIGOALFA"]])
    sv_code      = extraer_sv(codigo_alfa)
    nombre       = nombre_display(cliente_raw)

    items_extra = []  # ítems 3-9 (ABONO, CANON, COBRADOR, REMITOS, equipos…)

    for fila in filas:
        codigo_pro = str(fila[col["CODIGOPRO"]]).strip()
        desc       = str(fila[col["NOMBRE"]]).strip()
        precio     = float(fila[col["PRECIOFINA"]] or 0)
        cantidad   = float(fila[col["CANTIDAD"]] or 1)
        u_val      = float(fila[col["TIP_LETRA"]] or 0)

        # Fila de encabezado (U=5, S='9999', desc='SERVICIOS DEL MES...'): ya la generamos fija
        if u_val == 5 and codigo_pro == "9999":
            continue
        # Fila de código SV (U=1, S='9999', desc empieza con '(SV-'): ya la generamos fija
        if u_val == 1 and codigo_pro == "9999" and desc.startswith("("):
            continue
        # Remitos con precio 0: incluir sin importe
        if desc.startswith("REMITOS"):
            items_extra.append({"desc": desc, "cantidad": int(cantidad), "importe": 0.0})
            continue
        # Resto: incluir si tienen precio o descripción relevante
        if precio > 0 or codigo_pro not in ("9999",):
            items_extra.append({"desc": desc, "cantidad": int(cantidad), "importe": precio})

    return {
        "nombre":      nombre,
        "tipo":        "CONSUMIDOR FINAL",
        "direccion":   direccion,
        "sv_code":     sv_code,
        "codigo_alfa": codigo_alfa,
        "cliente_raw": cliente_raw,
        "mes":         mes,
        "anio":        anio,
        "items_extra": items_extra,
    }

def escribir_arca(clientes: list, mes: int, anio: int, salida: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Hoja1"

    for idx, c in enumerate(clientes):
        r = idx + 1
        mes_nombre = MESES[mes]
        sv          = c["sv_code"]
        nombre      = c["nombre"]
        cliente_raw = c["cliente_raw"]
        codigo_alfa = c["codigo_alfa"]
        items_extra = c["items_extra"]

        # Calcular importes primero para poder poner el total como valor (no fórmula)
        importes = []
        for item in items_extra:
            importes.append(aplicar_iva(item["importe"], item["desc"]) if item["importe"] > 0 else 0.0)
        total = round(sum(importes), 2)

        # Bloque fijo
        ws.cell(r, COL_NOMBRE).value = nombre
        ws.cell(r, COL_TIPO).value   = "CONSUMIDOR FINAL"
        ws.cell(r, COL_DIR).value    = None  # dirección vacía
        ws.cell(r, COL_TOTAL).value  = total  # valor calculado, no fórmula

        # Ítem 1: encabezado del mes (cols F=6, G=7, H=8, I=9)
        ws.cell(r, 6).value = 1.0
        ws.cell(r, 7).value = 1.0
        ws.cell(r, 8).value = f'SERVICIOS DEL MES "{mes_nombre}" DE {anio}'
        ws.cell(r, 9).value = 0.0

        # Ítem 2: código SV + nombre cliente (cols J=10, K=11, L=12, M=13)
        ws.cell(r, 10).value = 2.0
        ws.cell(r, 11).value = 1.0
        ws.cell(r, 12).value = f"({codigo_alfa.split('/')[0]}) {cliente_raw}"
        ws.cell(r, 13).value = 0.0

        # Ítems 3+ (sin límite)
        for i, item in enumerate(items_extra):
            base = 14 + i * 4
            ws.cell(r, base).value     = float(i + 3)
            ws.cell(r, base + 1).value = 1.0 if item["desc"] else None
            ws.cell(r, base + 2).value = item["desc"] or None
            ws.cell(r, base + 3).value = importes[i]

    wb.save(salida)
    print(f"OK Generado: {salida} ({len(clientes)} clientes)")


def main():
    if len(sys.argv) < 2:
        print("Uso: py scripts/convertir_reporte.py reporte.xlsx [mes] [anio] [salida.xlsx]")
        sys.exit(1)

    reporte = sys.argv[1]
    mes     = int(sys.argv[2]) if len(sys.argv) > 2 else None
    anio    = int(sys.argv[3]) if len(sys.argv) > 3 else None
    salida  = sys.argv[4] if len(sys.argv) > 4 else "FC_ARCA_SALIDA.xlsx"

    grupos, col = leer_reporte(reporte)

    # Detectar mes/año automáticamente desde el reporte si no se pasan
    if not mes or not anio:
        primera_fecha = None
        for filas in grupos.values():
            for fila in filas:
                fecha = fila[col["FECHA"]]
                if fecha:
                    primera_fecha = fecha
                    break
            if primera_fecha:
                break
        if primera_fecha:
            mes  = mes  or primera_fecha.month
            anio = anio or primera_fecha.year
        else:
            print("No se pudo detectar mes/año. Pasalos como argumento.")
            sys.exit(1)

    # Diccionario de direcciones (se completa cuando llegue el otro reporte)
    # Formato: { "SUOEM": "LIBERTAD N° 1450 SAN FRANCISCO", ... }
    DIRECCIONES = {}

    clientes_procesados = []
    for oid, filas in grupos.items():
        cliente_raw = str(filas[0][col["CLIENTE"]])
        dir_cliente = DIRECCIONES.get(cliente_raw, "")
        datos = procesar_cliente(filas, col, mes, anio, dir_cliente)
        clientes_procesados.append(datos)

    escribir_arca(clientes_procesados, mes, anio, salida)


if __name__ == "__main__":
    main()
